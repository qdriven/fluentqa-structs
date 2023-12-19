#!/usr/bin/env python
from typing import Any
from typing import List
from typing import Type

import csv

from pathlib import Path

import pyexcel

from openpyxl import load_workbook
from pydantic import BaseModel


__all__ = [
    "read_excel_to_objects",
    "write_objects_to_excel",
    "write_objects_to_csv",
    "read_csv_to_objects",
]

from qpystructs.models import BaseDataModel


def read_excel_to_objects(
    excel_path: str | Path,
    model: type[BaseModel] = BaseModel,
    sheet_index: int = 0,
    sheet_name: str = None,
    ignore_validate_errors: bool = False,
) -> list[Any]:
    wb = load_workbook(excel_path)
    if sheet_name is not None:
        sheet = wb.get_sheet_by_name(sheet_name)
    else:
        sheet = wb.worksheets[sheet_index]
    rows = sheet.iter_rows()
    headers = {th.value: i for i, th in enumerate(next(rows))}
    objects = []
    for row_index, raw_row in enumerate(rows, start=2):
        try:
            item = {key: v.value for key, v in zip(headers, raw_row) if key is not None}
            objects.append(model.parse_obj(item))
        except BaseException as e:
            if ignore_validate_errors:
                continue
            raise e
    return objects


def write_objects_to_excel(
    objects: list[BaseModel | dict | BaseDataModel], excel_path: str
):
    """
    write objects to excel or csv
    """
    if isinstance(objects[0], dict) or isinstance(objects[0], dict):
        pyexcel.get_sheet(records=objects).save_as(excel_path)
    elif isinstance(objects[0], BaseModel):
        pyexcel.get_sheet(
            records=[item.dict(by_alias=True) for item in objects]
        ).save_as(excel_path)
    else:
        raise NotImplementedError("Not Support Model type")


def write_objects_to_csv(csv_path: str, data: List[BaseModel]):
    out_csv = Path(csv_path)

    fields = list(data[0].__fields__)
    with out_csv.open("w") as out_fp:
        writer = csv.DictWriter(out_fp, fieldnames=fields)
        writer.writeheader()
        [writer.writerow(x.dict()) for x in data]


def read_csv_to_objects(
    csv_path: str, model: Type[BaseModel], ignore_validate_errors=False
):
    with open(csv_path, newline="") as f:
        reader = csv.reader(f, delimiter=":", quoting=csv.QUOTE_NONE)
        index = 0
        headers = []
        objects = []
        for row in reader:
            if index == 0:
                headers = row[0].split(",")
            else:
                try:
                    item = {}
                    for key, v in zip(headers, row[0].split(",")):
                        if len(v) > 0:
                            item[key] = v
                    objects.append(model.parse_obj(item))
                except BaseException as e:
                    if ignore_validate_errors:
                        continue
                    raise e
            index += 1

    return objects
